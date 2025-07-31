from fastapi import FastAPI, HTTPException, Depends, status, Request, File, UploadFile, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse
import pandas as pd
import numpy as np
from datetime import datetime, date
import os
import tempfile
import json
import uuid
from typing import List, Optional, Dict, Any
import uvicorn
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import jwt
import psycopg2
from psycopg2.extras import RealDictCursor
import zipfile
import io

app = FastAPI(title="DIAN Processing Service", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for now
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration
OUTPUT_FOLDER = "dian_reports"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# JWT configuration
JWT_SECRET = os.getenv("JWT_SECRET", "your-secret-key")
JWT_ALGORITHM = "HS256"

# Database configuration
DB_CONFIG = {
    "host": os.getenv("DB_HOST", "postgres"),
    "database": os.getenv("DB_NAME", "dian_saas"),
    "user": os.getenv("DB_USER", "admin"),
    "password": os.getenv("DB_PASSWORD", "admin123"),
    "port": os.getenv("DB_PORT", "5432")
}

# JWT Bearer token
security = HTTPBearer()

# Expected column mappings for DIAN compliance
DIAN_COLUMNS = {
    'FECHA': ['fecha', 'date', 'fecha_transaccion', 'fecha_documento'],
    'VALOR': ['valor', 'monto', 'amount', 'monetary_amount', 'valor_transaccion'],
    'CUENTA': ['cuenta', 'account', 'account_code', 'codigo_cuenta'],
    'NIT': ['nit', 'identificacion_tercero', 'third_party_id', 'nit_tercero'],
    'TIPO_DOCUMENTO': ['tipo_documento', 'document_type', 'tipo', 'documento']
}

@app.get("/")
async def root():
    """Root endpoint - Welcome to DIAN Processing Service"""
    return {
        "message": "Bienvenido a la API de DIAN Processing Service",
        "version": "1.0.0",
        "description": "Servicio para procesar archivos de contabilidad y generar reportes DIAN",
        "endpoints": {
            "health": "/health - Verificar estado del servicio",
            "docs": "/docs - Documentación de la API",
            "process_files": "/process_dian_files - Procesar archivos de contabilidad",
            "download_report": "/download_dian_report/{filename} - Descargar reportes generados"
        },
        "supported_formats": ["1001", "1002", "1003"],
        "supported_export_formats": ["excel", "pdf"],
        "supported_file_types": [".xlsx", ".xls", ".csv"]
    }

def get_db_connection():
    """Create database connection"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None

def verify_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """Verify JWT token"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token has expired"
        )
    except (jwt.InvalidTokenError, jwt.InvalidSignatureError, jwt.DecodeError):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Invalid token"
        )

def log_processing(user_id: str, processing_type: str, status: str, metadata: dict = None):
    """Log processing execution to database"""
    conn = get_db_connection()
    if not conn:
        return None
    
    try:
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                INSERT INTO workflow_logs (id, user_id, workflow_type, status, metadata, created_at)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                str(uuid.uuid4()),
                user_id,
                processing_type,
                status,
                json.dumps(metadata) if metadata else None,
                datetime.utcnow()
            ))
            conn.commit()
            result = cur.fetchone()
            return result['id'] if result else None
    except Exception as e:
        print(f"Database error: {e}")
        conn.rollback()
        return None
    finally:
        conn.close()

def normalize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """Normalize column names to standard DIAN format"""
    column_mapping = {}
    
    for standard_col, possible_names in DIAN_COLUMNS.items():
        for col in df.columns:
            if col.lower() in [name.lower() for name in possible_names]:
                column_mapping[col] = standard_col
                break
    
    df = df.rename(columns=column_mapping)
    return df

def parse_date(date_str: str) -> Optional[date]:
    """Parse date string in various Colombian formats"""
    if pd.isna(date_str) or str(date_str).strip() == '':
        return None
    
    date_str = str(date_str).strip()
    
    # Try different date formats
    date_formats = [
        '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%m-%y',
        '%Y/%m/%d', '%d/%m/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S'
    ]
    
    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    
    return None

def validate_and_clean_data(df: pd.DataFrame) -> pd.DataFrame:
    """Validate and clean the dataframe according to DIAN requirements"""
    # Normalize column names
    df = normalize_column_names(df)
    
    # Remove completely empty rows
    df = df.dropna(how='all')
    
    # Parse and validate dates
    if 'FECHA' in df.columns:
        df['FECHA_PARSED'] = df['FECHA'].apply(parse_date)
        # Remove rows with invalid dates
        df = df.dropna(subset=['FECHA_PARSED'])
    
    # Validate VALOR column
    if 'VALOR' in df.columns:
        # Convert to numeric, coerce errors to NaN
        df['VALOR'] = pd.to_numeric(df['VALOR'], errors='coerce')
        # Remove rows with invalid values
        df = df.dropna(subset=['VALOR'])
        # Remove zero or negative values (optional - depends on business logic)
        df = df[df['VALOR'] > 0]
    
    return df

def filter_by_date_range(df: pd.DataFrame, start_date: date, end_date: date) -> pd.DataFrame:
    """Filter dataframe by date range"""
    if 'FECHA_PARSED' not in df.columns:
        return df
    
    mask = (df['FECHA_PARSED'] >= start_date) & (df['FECHA_PARSED'] <= end_date)
    return df[mask]

def calculate_summaries(df: pd.DataFrame) -> Dict[str, Any]:
    """Calculate various summaries for DIAN reporting"""
    if df.empty:
        return {
            "total_amount": 0,
            "total_transactions": 0,
            "monthly_breakdown": {},
            "account_breakdown": {},
            "nit_breakdown": {},
            "document_type_breakdown": {}
        }
    
    # Check if VALOR column exists
    if 'VALOR' not in df.columns:
        raise KeyError("VALOR column not found after normalization")
    
    # Total summary
    total_amount = df['VALOR'].sum()
    total_transactions = len(df)
    
    # Monthly breakdown
    monthly_breakdown = {}
    if 'FECHA_PARSED' in df.columns:
        # Convert to datetime if it's not already
        if not pd.api.types.is_datetime64_any_dtype(df['FECHA_PARSED']):
            df['FECHA_PARSED'] = pd.to_datetime(df['FECHA_PARSED'])
        
        df['YEAR_MONTH'] = df['FECHA_PARSED'].dt.to_period('M')
        monthly_breakdown = df.groupby('YEAR_MONTH')['VALOR'].agg(['sum', 'count']).to_dict('index')
        # Convert Period objects to strings
        monthly_breakdown = {str(k): v for k, v in monthly_breakdown.items()}
    
    # Account breakdown
    account_breakdown = {}
    if 'CUENTA' in df.columns:
        account_breakdown = df.groupby('CUENTA')['VALOR'].agg(['sum', 'count']).to_dict('index')
    
    # NIT breakdown
    nit_breakdown = {}
    if 'NIT' in df.columns:
        nit_breakdown = df.groupby('NIT')['VALOR'].agg(['sum', 'count']).to_dict('index')
    
    # Document type breakdown
    document_type_breakdown = {}
    if 'TIPO_DOCUMENTO' in df.columns:
        document_type_breakdown = df.groupby('TIPO_DOCUMENTO')['VALOR'].agg(['sum', 'count']).to_dict('index')
    
    return {
        "total_amount": float(total_amount),
        "total_transactions": int(total_transactions),
        "monthly_breakdown": monthly_breakdown,
        "account_breakdown": account_breakdown,
        "nit_breakdown": nit_breakdown,
        "document_type_breakdown": document_type_breakdown
    }

def generate_dian_excel_report(df: pd.DataFrame, summaries: Dict[str, Any], 
                             start_date: date, end_date: date, formato: str = "1001") -> str:
    """Generate DIAN-compliant Excel report"""
    filename = f"dian_report_{formato}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Summary sheet
        summary_data = []
        summary_data.append(["DIAN Report Summary", ""])
        summary_data.append(["Formato", formato])
        summary_data.append(["Date Range", f"{start_date} to {end_date}"])
        summary_data.append(["Total Amount", f"${summaries['total_amount']:,.2f}"])
        summary_data.append(["Total Transactions", summaries['total_transactions']])
        summary_data.append(["", ""])
        
        # Monthly breakdown
        if summaries['monthly_breakdown']:
            summary_data.append(["Monthly Breakdown", ""])
            summary_data.append(["Month", "Total Amount", "Transaction Count"])
            for month, data in summaries['monthly_breakdown'].items():
                summary_data.append([month, f"${data['sum']:,.2f}", data['count']])
            summary_data.append(["", ""])
        
        # Account breakdown
        if summaries['account_breakdown']:
            summary_data.append(["Account Breakdown", ""])
            summary_data.append(["Account", "Total Amount", "Transaction Count"])
            for account, data in summaries['account_breakdown'].items():
                summary_data.append([account, f"${data['sum']:,.2f}", data['count']])
            summary_data.append(["", ""])
        
        # NIT breakdown
        if summaries['nit_breakdown']:
            summary_data.append(["NIT Breakdown", ""])
            summary_data.append(["NIT", "Total Amount", "Transaction Count"])
            for nit, data in summaries['nit_breakdown'].items():
                summary_data.append([nit, f"${data['sum']:,.2f}", data['count']])
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
        
        # Transactions sheet
        if not df.empty:
            # Select relevant columns for output
            output_columns = ['FECHA_PARSED', 'VALOR', 'CUENTA', 'NIT', 'TIPO_DOCUMENTO']
            output_df = df[output_columns].copy()
            output_df.columns = ['FECHA', 'VALOR', 'CUENTA', 'NIT', 'TIPO_DOCUMENTO']
            output_df['FECHA'] = output_df['FECHA'].dt.strftime('%d/%m/%Y')
            output_df.to_excel(writer, sheet_name='Transactions', index=False)
        
        # Format the summary sheet
        workbook = writer.book
        worksheet = writer.sheets['Summary']
        
        # Apply formatting
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value and 'DIAN Report Summary' in str(cell.value):
                    cell.font = Font(bold=True, size=14)
                elif cell.value and any(keyword in str(cell.value) for keyword in ['Breakdown', 'Summary']):
                    cell.font = Font(bold=True)
    
    return filepath

def generate_dian_pdf_report(df: pd.DataFrame, summaries: Dict[str, Any], 
                           start_date: date, end_date: date, formato: str = "1001") -> str:
    """Generate DIAN-compliant PDF report"""
    filename = f"dian_report_{formato}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    
    pdf = FPDF()
    pdf.add_page()
    
    # Header
    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 15, txt=f"DIAN Report - Formato {formato}", ln=True, align="C")
    pdf.cell(200, 10, txt="Dirección de Impuestos y Aduanas Nacionales", ln=True, align="C")
    pdf.ln(10)
    
    # Report metadata
    pdf.set_font("Arial", "B", 12)
    pdf.cell(200, 10, txt=f"Date Range: {start_date} to {end_date}", ln=True)
    pdf.cell(200, 10, txt=f"Total Amount: ${summaries['total_amount']:,.2f}", ln=True)
    pdf.cell(200, 10, txt=f"Total Transactions: {summaries['total_transactions']}", ln=True)
    pdf.ln(10)
    
    # Monthly breakdown
    if summaries['monthly_breakdown']:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(200, 10, txt="Monthly Breakdown", ln=True)
        pdf.set_font("Arial", "", 10)
        
        for month, data in list(summaries['monthly_breakdown'].items())[:12]:  # Limit to 12 months
            pdf.cell(200, 8, txt=f"{month}: ${data['sum']:,.2f} ({data['count']} transactions)", ln=True)
        pdf.ln(5)
    
    # Account breakdown
    if summaries['account_breakdown']:
        pdf.set_font("Arial", "B", 12)
        pdf.cell(200, 10, txt="Account Breakdown", ln=True)
        pdf.set_font("Arial", "", 10)
        
        for account, data in list(summaries['account_breakdown'].items())[:10]:  # Limit to 10 accounts
            pdf.cell(200, 8, txt=f"{account}: ${data['sum']:,.2f} ({data['count']} transactions)", ln=True)
        pdf.ln(5)
    
    # Footer
    pdf.ln(10)
    pdf.set_font("Arial", "I", 10)
    pdf.cell(200, 10, txt=f"Generated on: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", ln=True)
    pdf.cell(200, 10, txt="Document generated automatically by DIAN Processing Service", ln=True)
    
    pdf.output(filepath)
    return filepath

@app.post("/process_dian_files")
async def process_dian_files(
    request: Request,
    start_date: str = Form(...),
    end_date: str = Form(...),
    formato: str = Form("1001"),
    export_format: str = Form("excel"),
    token_payload: dict = Depends(verify_token)
):
    """Process multiple files for DIAN reporting"""
    user_id = token_payload.get("sub", "unknown")
    processing_id = log_processing(user_id, "dian_processing", "started")
    
    try:
        print(f"Processing request for user {user_id}")
        print(f"Date range: {start_date} to {end_date}")
        print(f"Formato: {formato}, Export format: {export_format}")
        
        # Parse dates
        start_date_parsed = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date_parsed = datetime.strptime(end_date, "%Y-%m-%d").date()
        
        # Get uploaded files
        form_data = await request.form()
        files = []
        
        print(f"Form data keys: {list(form_data.keys())}")
        
        # Handle multiple file uploads
        for key, value in form_data.items():
            print(f"Processing form field: {key}, type: {type(value)}")
            if hasattr(value, 'filename') and hasattr(value, 'file'):
                files.append(value)
                print(f"Added file: {value.filename}")
            else:
                print(f"Skipping non-UploadFile: {key}")
        
        # Also check for single file upload with key 'file'
        if 'file' in form_data and hasattr(form_data['file'], 'filename'):
            if form_data['file'] not in files:
                files.append(form_data['file'])
                print(f"Added file from 'file' key: {form_data['file'].filename}")
            else:
                print(f"File already in list: {form_data['file'].filename}")
        else:
            print(f"File key not found or not UploadFile: {type(form_data.get('file'))}")
        
        print(f"Files list length: {len(files)}")
        for i, f in enumerate(files):
            print(f"  File {i}: {f.filename}")
        
        if not files:
            error_msg = "No files provided"
            print(f"Error: {error_msg}")
            log_processing(user_id, "dian_processing", "failed", {"error": error_msg})
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        
        print(f"Processing {len(files)} files")
        
        # Process all files
        all_data = []
        file_summaries = []
        
        for file in files:
            try:
                print(f"Processing file: {file.filename}")
                
                # Save uploaded file to temporary location
                temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1])
                file.file.seek(0)
                temp_file.write(file.file.read())
                temp_file.close()
                
                # Read file based on extension
                if file.filename.lower().endswith('.csv'):
                    df = pd.read_csv(temp_file.name)
                elif file.filename.lower().endswith(('.xlsx', '.xls')):
                    df = pd.read_excel(temp_file.name)
                else:
                    print(f"Skipping unsupported file: {file.filename}")
                    os.unlink(temp_file.name)
                    continue
                
                # Clean up temporary file
                os.unlink(temp_file.name)
                
                print(f"File {file.filename}: {len(df)} rows, columns: {list(df.columns)}")
                
                # Validate and clean data
                df = validate_and_clean_data(df)
                print(f"After validation: {len(df)} rows")
                
                # Filter by date range
                df_filtered = filter_by_date_range(df, start_date_parsed, end_date_parsed)
                print(f"After date filtering: {len(df_filtered)} rows")
                
                # Add file identifier
                df_filtered['SOURCE_FILE'] = file.filename
                
                all_data.append(df_filtered)
                
                # Calculate file summary
                file_summary = calculate_summaries(df_filtered)
                file_summaries.append({
                    "filename": file.filename,
                    "total_rows": len(df),
                    "filtered_rows": len(df_filtered),
                    "total_amount": file_summary["total_amount"]
                })
                
                print(f"File {file.filename} processed successfully")
                
            except Exception as e:
                error_msg = f"Error processing file {file.filename}: {str(e)}"
                print(error_msg)
                continue
        
        if not all_data:
            error_msg = "No valid data found in uploaded files"
            print(f"Error: {error_msg}")
            log_processing(user_id, "dian_processing", "failed", {"error": error_msg})
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=error_msg
            )
        
        print(f"Combining {len(all_data)} dataframes")
        
        # Combine all data
        combined_df = pd.concat(all_data, ignore_index=True)
        print(f"Combined dataframe: {len(combined_df)} rows")
        
        # Calculate overall summaries
        overall_summaries = calculate_summaries(combined_df)
        print(f"Overall summaries calculated: {overall_summaries['total_transactions']} transactions, ${overall_summaries['total_amount']:,.2f}")
        
        # Generate report
        print(f"Generating {export_format} report with formato {formato}")
        if export_format.lower() == "pdf":
            report_path = generate_dian_pdf_report(combined_df, overall_summaries, start_date_parsed, end_date_parsed, formato)
        else:
            report_path = generate_dian_excel_report(combined_df, overall_summaries, start_date_parsed, end_date_parsed, formato)
        
        print(f"Report generated: {report_path}")
        
        # Log successful processing
        log_processing(user_id, "dian_processing", "completed", {
            "files_processed": len(files),
            "total_transactions": overall_summaries["total_transactions"],
            "total_amount": overall_summaries["total_amount"],
            "formato": formato,
            "export_format": export_format
        })
        
        return {
            "success": True,
            "processing_id": processing_id,
            "file_summaries": file_summaries,
            "overall_summaries": overall_summaries,
            "report_filename": os.path.basename(report_path),
            "download_url": f"/download_dian_report/{os.path.basename(report_path)}",
            "message": f"Successfully processed {len(files)} files"
        }
        
    except Exception as e:
        error_msg = f"Processing failed: {str(e)}"
        print(f"Exception in process_dian_files: {error_msg}")
        import traceback
        traceback.print_exc()
        log_processing(user_id, "dian_processing", "failed", {"error": str(e)})
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=error_msg
        )

@app.get("/download_dian_report/{filename}")
async def download_dian_report(filename: str, token_payload: dict = Depends(verify_token)):
    """Download generated DIAN report"""
    filepath = os.path.join(OUTPUT_FOLDER, filename)
    
    if not os.path.exists(filepath):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Report file not found"
        )
    
    return FileResponse(
        filepath,
        media_type="application/octet-stream",
        filename=filename,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "service": "dian-processing-service"}

if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 8003))
    uvicorn.run(app, host="0.0.0.0", port=port) 